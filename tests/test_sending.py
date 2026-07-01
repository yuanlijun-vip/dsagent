import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

from refund_recovery.sending import SEND_STATUS_RESERVED, STATUS_PENDING, STATUS_SENDING
from refund_recovery.sending import next_send_candidates, release_reservations, reserve_candidates


class SendingTests(unittest.TestCase):
    def make_workbook(self) -> Path:
        tmp = tempfile.TemporaryDirectory()
        self.addCleanup(tmp.cleanup)
        path = Path(tmp.name) / "followup.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "跟进明细"
        ws.append(["月份", "店铺名", "订单号", "旺旺ID", "退款申请时间", "订单状态", "", "", "", "处理状态", "", "导出文件名", "", "更新时间", "发送状态", "发送时间", "发送内容版本", "发送详情"])
        ws.append(["2026-06", "Shop A", "1001", "buyer-a", "", "", "", "", "", STATUS_PENDING, "", "a.xlsx", "", "", "", "", "", ""])
        ws.append(["2026-06", "Shop A", "1002", "buyer-b", "", "", "", "", "", "已发送", "", "a.xlsx", "", "", "已发送", "", "", ""])
        ws.append(["2026-06", "Shop B", "1003", "buyer-c", "", "", "", "", "", STATUS_PENDING, "", "b.xlsx", "", "", "", "", "", ""])
        wb.save(path)
        return path

    def test_next_send_candidates_filters_pending_unsent(self):
        path = self.make_workbook()

        candidates = next_send_candidates(path, limit=5, shop="Shop A")

        self.assertEqual([candidate.order_id for candidate in candidates], ["1001"])

    def test_reserve_and_release_candidates(self):
        path = self.make_workbook()
        [candidate] = next_send_candidates(path, limit=1)

        reserve_candidates(path, [candidate], reserved_at="2026-07-01 13:00:00")

        wb = load_workbook(path, data_only=True)
        ws = wb["跟进明细"]
        self.assertEqual(ws.cell(candidate.row, 10).value, STATUS_SENDING)
        self.assertEqual(ws.cell(candidate.row, 15).value, SEND_STATUS_RESERVED)

        release_reservations(path, [candidate.row])
        wb = load_workbook(path, data_only=True)
        ws = wb["跟进明细"]
        self.assertEqual(ws.cell(candidate.row, 10).value, STATUS_PENDING)
        self.assertIsNone(ws.cell(candidate.row, 15).value)


if __name__ == "__main__":
    unittest.main()
