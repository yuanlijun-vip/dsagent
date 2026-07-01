import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook

from refund_recovery.closure import STATUS_PENDING, STATUS_SENT, STATUS_SEND_ERROR, STATUS_PARTIAL_SENT
from refund_recovery.closure import send_summary_by_shop, summary_row_values, update_local_summary


class ClosureTests(unittest.TestCase):
    def test_send_summary_by_shop_counts_process_and_send_status(self):
        wb = Workbook()
        ws = wb.active
        ws.append([""] * 18)
        ws.append(["", "Shop A", "", "", "", "", "", "", "", STATUS_PENDING, "", "", "", "", "", "", "", ""])
        ws.append(["", "Shop A", "", "", "", "", "", "", "", STATUS_SENT, "", "", "", "", STATUS_SENT, "", "", ""])
        ws.append(
            ["", "Shop A", "", "", "", "", "", "", "", STATUS_SEND_ERROR, "", "", "", "", STATUS_PARTIAL_SENT, "", "", ""]
        )

        summary = send_summary_by_shop(ws)

        self.assertEqual(summary_row_values(summary["Shop A"]), [1, 1, 1, 1])

    def test_update_local_summary_writes_existing_summary_rows(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "followup.xlsx"
            wb = Workbook()
            detail = wb.active
            detail.title = "跟进明细"
            detail.append([""] * 18)
            detail.append(["", "Shop A", "", "", "", "", "", "", "", STATUS_SENT, "", "", "", "", STATUS_SENT, "", "", ""])
            summary = wb.create_sheet("汇总")
            summary.append(["店铺名"])
            summary.append(["Shop A"])
            summary.append(["合计"])
            wb.save(path)

            update_local_summary(path)

            loaded = Workbook()
            loaded = __import__("openpyxl").load_workbook(path, data_only=True)
            ws = loaded["汇总"]
            self.assertEqual([ws.cell(1, col).value for col in range(9, 13)], ["发送后待发送", "已发送", "发送异常", "部分发送"])
            self.assertEqual([ws.cell(2, col).value for col in range(9, 13)], [0, 1, 0, 0])
            self.assertEqual([ws.cell(3, col).value for col in range(9, 13)], [0, 1, 0, 0])


if __name__ == "__main__":
    unittest.main()
