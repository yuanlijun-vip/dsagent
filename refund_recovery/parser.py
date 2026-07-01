from __future__ import annotations

import csv
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook

from .models import RefundOrder


DEFAULT_COLUMN_ALIASES = {
    "order_id": ["订单号", "主订单编号", "子订单编号"],
    "buyer_id": ["旺旺ID", "买家旺旺", "买家昵称"],
    "refund_time": ["退款申请时间", "申请时间", "售后申请时间"],
    "order_status": ["订单状态", "交易状态"],
    "today_chatted": ["今日是否已对话", "是否今日已对话"],
    "has_unfinished_order": ["是否未完结订单", "存在未完结订单"],
}


class ExportParseError(RuntimeError):
    pass


def iter_export_files(export_dir: Path) -> list[Path]:
    if not export_dir.exists():
        return []
    return sorted(
        [
            path
            for path in export_dir.iterdir()
            if path.is_file() and path.suffix.lower() in {".xlsx", ".csv"}
        ]
    )


def parse_export_file(path: Path, shop_name: str, aliases: dict[str, list[str]] | None = None) -> list[RefundOrder]:
    aliases = {**DEFAULT_COLUMN_ALIASES, **(aliases or {})}
    rows = _read_rows(path)
    if not rows:
        return []
    header_index = _find_header_index(rows, aliases)
    headers = [str(cell).strip() for cell in rows[header_index]]
    mapping = _resolve_columns(headers, aliases)
    orders: list[RefundOrder] = []
    for row in rows[header_index + 1 :]:
        if not any(str(cell).strip() for cell in row):
            continue
        orders.append(
            RefundOrder(
                shop_name=shop_name,
                order_id=_cell(row, mapping.get("order_id", -1)),
                buyer_id=_cell(row, mapping.get("buyer_id", -1)),
                refund_time=_cell(row, mapping.get("refund_time", -1)),
                order_status=_cell(row, mapping.get("order_status", -1)),
                source_file=path.name,
                today_chatted=_truthy(_cell(row, mapping.get("today_chatted", -1))),
                has_unfinished_order=_truthy(_cell(row, mapping.get("has_unfinished_order", -1))),
            )
        )
    return orders


def _read_rows(path: Path) -> list[list[str]]:
    suffix = path.suffix.lower()
    if suffix == ".xlsx":
        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook.active
        return [[_normalize(cell) for cell in row] for row in sheet.iter_rows(values_only=True)]
    if suffix == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            return [[_normalize(cell) for cell in row] for row in csv.reader(handle)]
    raise ExportParseError(f"Unsupported export file: {path}")


def _find_header_index(rows: list[list[str]], aliases: dict[str, list[str]]) -> int:
    required_aliases = aliases["order_id"]
    for index, row in enumerate(rows[:20]):
        values = {str(cell).strip() for cell in row}
        if any(alias in values for alias in required_aliases):
            return index
    raise ExportParseError("Cannot find header row with order id column")


def _resolve_columns(headers: list[str], aliases: dict[str, list[str]]) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for field, names in aliases.items():
        for name in names:
            if name in headers:
                mapping[field] = headers.index(name)
                break
    if "order_id" not in mapping:
        raise ExportParseError("Missing required order id column")
    return mapping


def _cell(row: Iterable[str], index: int) -> str:
    row_list = list(row)
    if index < 0 or index >= len(row_list):
        return ""
    return str(row_list[index]).strip()


def _normalize(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _truthy(value: str) -> bool:
    return value.strip().lower() in {"是", "有", "true", "yes", "y", "1"}
