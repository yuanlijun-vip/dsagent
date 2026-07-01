from __future__ import annotations

from dataclasses import asdict, dataclass
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

from .closure import DETAIL_SHEET, cell_text


STATUS_PENDING = "\u5f85\u53d1\u9001"
STATUS_SENDING = "\u53d1\u9001\u4e2d"
SEND_STATUS_RESERVED = "\u5df2\u9501\u5b9a"


@dataclass(frozen=True)
class SendCandidate:
    row: int
    month: str
    shop: str
    order_id: str
    buyer_id: str
    refund_time: str
    order_status: str
    source_file: str

    def to_dict(self) -> dict[str, object]:
        return asdict(self)


def is_available_for_send(process_status: str, send_status: str) -> bool:
    return process_status == STATUS_PENDING and not send_status


def next_send_candidates(workbook_path: Path, limit: int = 1, shop: str | None = None) -> list[SendCandidate]:
    wb = load_workbook(workbook_path, data_only=True)
    ws = wb[DETAIL_SHEET] if DETAIL_SHEET in wb.sheetnames else wb.worksheets[0]
    candidates: list[SendCandidate] = []
    for row_index in range(2, ws.max_row + 1):
        row_shop = cell_text(ws.cell(row_index, 2).value).strip()
        if shop and row_shop != shop:
            continue
        process_status = cell_text(ws.cell(row_index, 10).value).strip()
        send_status = cell_text(ws.cell(row_index, 15).value).strip()
        if not is_available_for_send(process_status, send_status):
            continue
        candidates.append(
            SendCandidate(
                row=row_index,
                month=cell_text(ws.cell(row_index, 1).value),
                shop=row_shop,
                order_id=cell_text(ws.cell(row_index, 3).value),
                buyer_id=cell_text(ws.cell(row_index, 4).value),
                refund_time=cell_text(ws.cell(row_index, 5).value),
                order_status=cell_text(ws.cell(row_index, 6).value),
                source_file=cell_text(ws.cell(row_index, 12).value),
            )
        )
        if len(candidates) >= limit:
            break
    return candidates


def reserve_candidates(workbook_path: Path, candidates: list[SendCandidate], reserved_at: str | None = None) -> None:
    if not candidates:
        return
    reserved_at = reserved_at or datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    wb = load_workbook(workbook_path)
    ws = wb[DETAIL_SHEET] if DETAIL_SHEET in wb.sheetnames else wb.worksheets[0]
    for candidate in candidates:
        current_process = cell_text(ws.cell(candidate.row, 10).value).strip()
        current_send = cell_text(ws.cell(candidate.row, 15).value).strip()
        if not is_available_for_send(current_process, current_send):
            raise ValueError(f"Row {candidate.row} is no longer available for send")
        ws.cell(candidate.row, 10, STATUS_SENDING)
        ws.cell(candidate.row, 14, reserved_at)
        ws.cell(candidate.row, 15, SEND_STATUS_RESERVED)
        ws.cell(candidate.row, 16, reserved_at)
        ws.cell(candidate.row, 18, "\u5df2\u9501\u5b9a\u5f85\u81ea\u52a8\u53d1\u9001")
    wb.save(workbook_path)


def release_reservations(workbook_path: Path, rows: list[int]) -> None:
    if not rows:
        return
    wb = load_workbook(workbook_path)
    ws = wb[DETAIL_SHEET] if DETAIL_SHEET in wb.sheetnames else wb.worksheets[0]
    for row in rows:
        if cell_text(ws.cell(row, 10).value).strip() == STATUS_SENDING:
            ws.cell(row, 10, STATUS_PENDING)
        if cell_text(ws.cell(row, 15).value).strip() == SEND_STATUS_RESERVED:
            ws.cell(row, 15).value = None
            ws.cell(row, 16).value = None
            ws.cell(row, 18).value = None
    wb.save(workbook_path)
