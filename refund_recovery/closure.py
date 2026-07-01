from __future__ import annotations

from collections import Counter, defaultdict
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

from .dingtalk import DingTalkClient


DETAIL_SHEET = "\u8ddf\u8fdb\u660e\u7ec6"
SUMMARY_SHEET = "\u6c47\u603b"
STATUS_PENDING = "\u5f85\u53d1\u9001"
STATUS_SENT = "\u5df2\u53d1\u9001"
STATUS_SEND_ERROR = "\u53d1\u9001\u5f02\u5e38"
STATUS_PARTIAL_SENT = "\u90e8\u5206\u53d1\u9001"
TOTAL_LABEL = "\u5408\u8ba1"

SUMMARY_HEADERS = [
    "\u53d1\u9001\u540e\u5f85\u53d1\u9001",
    "\u5df2\u53d1\u9001",
    "\u53d1\u9001\u5f02\u5e38",
    "\u90e8\u5206\u53d1\u9001",
]


@dataclass(frozen=True)
class ClosureSyncResult:
    detail_rows_synced: int
    summary_rows_synced: int


def cell_text(value: object) -> str:
    return "" if value is None else str(value)


def load_detail_sheet(workbook_path: Path):
    wb = load_workbook(workbook_path)
    ws = wb[DETAIL_SHEET] if DETAIL_SHEET in wb.sheetnames else wb.worksheets[0]
    return wb, ws


def send_summary_by_shop(ws) -> dict[str, Counter]:
    summary: dict[str, Counter] = defaultdict(Counter)
    for row_index in range(2, ws.max_row + 1):
        shop = cell_text(ws.cell(row_index, 2).value).strip()
        if not shop:
            continue
        process_status = cell_text(ws.cell(row_index, 10).value).strip()
        send_status = cell_text(ws.cell(row_index, 15).value).strip()
        summary[shop]["total"] += 1
        if process_status:
            summary[shop][f"process:{process_status}"] += 1
        if send_status:
            summary[shop][f"send:{send_status}"] += 1
    return summary


def summary_row_values(counter: Counter) -> list[int]:
    return [
        counter.get(f"process:{STATUS_PENDING}", 0),
        counter.get(f"send:{STATUS_SENT}", 0),
        counter.get(f"process:{STATUS_SEND_ERROR}", 0),
        counter.get(f"send:{STATUS_PARTIAL_SENT}", 0),
    ]


def update_local_summary(workbook_path: Path) -> list[tuple[int, list[int]]]:
    wb, detail_ws = load_detail_sheet(workbook_path)
    summary_ws = wb[SUMMARY_SHEET] if SUMMARY_SHEET in wb.sheetnames else wb.create_sheet(SUMMARY_SHEET)
    by_shop = send_summary_by_shop(detail_ws)
    total = sum((Counter(value) for value in by_shop.values()), Counter())

    for offset, header in enumerate(SUMMARY_HEADERS, start=9):
        summary_ws.cell(1, offset, header)

    updated_rows: list[tuple[int, list[int]]] = []
    for row_index in range(2, summary_ws.max_row + 1):
        shop = cell_text(summary_ws.cell(row_index, 1).value).strip()
        if not shop:
            continue
        values = summary_row_values(total if shop == TOTAL_LABEL else by_shop.get(shop, Counter()))
        for offset, value in enumerate(values, start=9):
            summary_ws.cell(row_index, offset, value)
        updated_rows.append((row_index, values))

    wb.save(workbook_path)
    return updated_rows


def detail_sync_values(workbook_path: Path, row_indexes: list[int]) -> list[tuple[int, list[str]]]:
    wb = load_workbook(workbook_path, data_only=True)
    ws = wb[DETAIL_SHEET] if DETAIL_SHEET in wb.sheetnames else wb.worksheets[0]
    rows: list[tuple[int, list[str]]] = []
    for row_index in row_indexes:
        rows.append((row_index, [cell_text(ws.cell(row_index, column).value) for column in range(10, 19)]))
    return rows


def sync_closure_to_dingtalk(
    workbook_path: Path,
    dingtalk: DingTalkClient,
    node: str,
    detail_sheet_id: str,
    summary_sheet_id: str | None,
    detail_row_indexes: list[int],
) -> ClosureSyncResult:
    detail_count = 0
    for row_index, values in detail_sync_values(workbook_path, detail_row_indexes):
        dingtalk.update_range(node, detail_sheet_id, f"J{row_index}:R{row_index}", [values])
        detail_count += 1

    summary_count = 0
    if summary_sheet_id:
        summary_rows = update_local_summary(workbook_path)
        dingtalk.update_range(node, summary_sheet_id, "I1:L1", [SUMMARY_HEADERS])
        for row_index, values in summary_rows:
            dingtalk.update_range(node, summary_sheet_id, f"I{row_index}:L{row_index}", [values])
            summary_count += 1

    return ClosureSyncResult(detail_rows_synced=detail_count, summary_rows_synced=summary_count)
