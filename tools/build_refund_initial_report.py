from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook


ROOT = Path(__file__).resolve().parents[1]
EXPORT_DIR = ROOT / "exports"
OUT_FILE = ROOT / "exports" / "refund_initial_report_2026-06.xlsx"

HEADERS = [
    "月份",
    "店铺名",
    "订单号",
    "旺旺ID",
    "退款申请时间",
    "订单状态",
    "是否未完结订单",
    "今日是否已对话",
    "是否当月重复",
    "处理状态",
    "跳过原因",
    "导出文件名",
    "创建时间",
    "更新时间",
]


def cell(row: tuple[object, ...], headers: list[str], name: str) -> str:
    if name not in headers:
        return ""
    value = row[headers.index(name)]
    return "" if value is None else str(value).strip()


def shop_from_filename(path: Path) -> str:
    return path.name.split("_未发货退款_", 1)[0]


def main() -> None:
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = "跟进明细"
    ws_out.append(HEADERS)

    seen_order_keys: set[tuple[str, str]] = set()
    for path in sorted(EXPORT_DIR.glob("*_未发货退款_20260629-20260630.xlsx")):
        shop = shop_from_filename(path)
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        headers = ["" if value is None else str(value).strip() for value in next(rows)]
        for row in rows:
            order_id = cell(row, headers, "订单编号")
            if not order_id:
                continue
            key = (shop, order_id)
            if key in seen_order_keys:
                continue
            seen_order_keys.add(key)
            refund_time = cell(row, headers, "退款申请时间")
            order_status = cell(row, headers, "退款状态") or cell(row, headers, "货物状态")
            ws_out.append(
                [
                    "2026-06",
                    shop,
                    order_id,
                    "",
                    refund_time,
                    order_status,
                    "",
                    "",
                    "",
                    "待补旺旺ID",
                    "退款导出表不含旺旺ID，需从订单/千牛补齐",
                    path.name,
                    now,
                    now,
                ]
            )

    ws_out.freeze_panes = "A2"
    for column in ws_out.columns:
        letter = column[0].column_letter
        width = min(max(len(str(cell.value or "")) for cell in column) + 2, 42)
        ws_out.column_dimensions[letter].width = width

    wb_out.save(OUT_FILE)
    print(OUT_FILE)
    print(ws_out.max_row - 1)


if __name__ == "__main__":
    main()
