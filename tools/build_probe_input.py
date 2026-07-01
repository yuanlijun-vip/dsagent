from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
INITIAL_REPORT = ROOT / "exports" / "refund_initial_report_2026-06.xlsx"
OUT_FILE = ROOT / "exports" / "refund_probe_input.json"


def main() -> None:
    workbook = load_workbook(INITIAL_REPORT, read_only=True, data_only=True)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    headers = [str(value or "").strip() for value in next(rows)]
    shop_idx = headers.index("店铺名")
    order_idx = headers.index("订单号")
    refund_idx = headers.index("退款申请时间")
    source_idx = headers.index("导出文件名")

    items = []
    for row_number, row in enumerate(rows, start=2):
        shop = str(row[shop_idx] or "").strip()
        order_id = str(row[order_idx] or "").strip()
        if not shop or not order_id:
            continue
        items.append(
            {
                "row": row_number,
                "shop": shop,
                "order_id": order_id,
                "refund_time": str(row[refund_idx] or "").strip(),
                "source_file": str(row[source_idx] or "").strip(),
            }
        )

    OUT_FILE.write_text(json.dumps(items, ensure_ascii=False, indent=2), encoding="utf-8")
    print(OUT_FILE)
    print(len(items))


if __name__ == "__main__":
    main()
