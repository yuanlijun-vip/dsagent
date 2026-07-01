from __future__ import annotations

import json
from collections import defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


ROOT = Path(__file__).resolve().parents[1]
EXPORT_DIR = ROOT / "exports"
MONTH = "2026-06"
INITIAL_FILE = EXPORT_DIR / f"refund_initial_report_{MONTH}.xlsx"
OUT_FILE = EXPORT_DIR / f"refund_followup_{MONTH}.xlsx"

HEADERS = [
    "月份",
    "店铺名",
    "订单号",
    "旺旺ID",
    "退款申请时间",
    "订单状态",
    "是否未完结订单",
    "今日是否已对话",
    "是否当月重复",
    "处理状态",
    "跳过原因",
    "导出文件名",
    "创建时间",
    "更新时间",
]

SHOP_ORDER = ["艺颂旗舰店", "科润旗舰店", "朗域轩品旗舰店", "领域家具", "yuruifeng1009"]


def normalize(value: object) -> str:
    return "" if value is None else str(value).strip()


def load_initial_rows() -> list[dict[str, str]]:
    wb = load_workbook(INITIAL_FILE, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    headers = [normalize(value) for value in next(rows)]
    result: list[dict[str, str]] = []
    for excel_row, values in enumerate(rows, start=2):
        data = {header: normalize(values[index]) for index, header in enumerate(headers)}
        data["_excel_row"] = str(excel_row)
        if data.get("店铺名") and data.get("订单号"):
            result.append(data)
    return result


def load_probe_rows() -> dict[tuple[str, str], dict[str, object]]:
    probe_by_key: dict[tuple[str, str], dict[str, object]] = {}
    for path in EXPORT_DIR.glob("refund_probe_results_*.json"):
        if path.name == "refund_probe_results_sample.json":
            continue
        rows = json.loads(path.read_text(encoding="utf-8"))
        for row in rows:
            key = (normalize(row.get("shop")), normalize(row.get("order_id")))
            if key[0] and key[1]:
                probe_by_key[key] = row
    return probe_by_key


def bool_cn(value: bool) -> str:
    return "是" if value else "否"


def status_for(
    buyer_id: str,
    today_chatted: bool,
    is_duplicate: bool,
    has_unfinished_order: bool,
    probe_error: str,
) -> tuple[str, str]:
    reasons: list[str] = []
    if not buyer_id:
        reasons.append("未找到旺旺ID")
    if today_chatted:
        reasons.append("今日已对话")
    if is_duplicate:
        reasons.append("当月重复")
    if has_unfinished_order:
        reasons.append("存在未完结订单")
    if probe_error and probe_error not in reasons:
        reasons.append(probe_error)

    if not buyer_id:
        return "异常-未找到旺旺ID", "；".join(reasons)
    if today_chatted or is_duplicate or has_unfinished_order:
        return "跳过", "；".join(reasons)
    return "待发送", ""


def main() -> None:
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    initial_rows = load_initial_rows()
    probe_by_key = load_probe_rows()

    shop_rank = {shop: index for index, shop in enumerate(SHOP_ORDER)}
    initial_rows.sort(key=lambda row: (shop_rank.get(row["店铺名"], 99), int(row["_excel_row"])))

    seen_buyer_by_shop: defaultdict[tuple[str, str], int] = defaultdict(int)

    wb = Workbook()
    ws = wb.active
    ws.title = "跟进明细"
    ws.append(HEADERS)

    summary = defaultdict(int)
    per_shop = defaultdict(lambda: defaultdict(int))

    for row in initial_rows:
        shop = row["店铺名"]
        order_id = row["订单号"]
        probe = probe_by_key.get((shop, order_id), {})

        buyer_id = normalize(probe.get("buyer_id"))
        today_chatted = bool(probe.get("today_chatted"))
        unfinished_count = int(probe.get("unfinished_count") or 0)
        pending_ship_count = int(probe.get("pending_ship_count") or 0)
        has_unfinished_order = bool(probe.get("has_unfinished_order")) or unfinished_count > 0 or pending_ship_count > 0
        probe_error = normalize(probe.get("error"))

        duplicate_key = (shop, buyer_id)
        is_duplicate = False
        if buyer_id:
            seen_buyer_by_shop[duplicate_key] += 1
            is_duplicate = seen_buyer_by_shop[duplicate_key] > 1

        status, reason = status_for(
            buyer_id=buyer_id,
            today_chatted=today_chatted,
            is_duplicate=is_duplicate,
            has_unfinished_order=has_unfinished_order,
            probe_error=probe_error,
        )

        ws.append(
            [
                MONTH,
                shop,
                order_id,
                buyer_id,
                row.get("退款申请时间", ""),
                row.get("订单状态", ""),
                bool_cn(has_unfinished_order),
                bool_cn(today_chatted),
                bool_cn(is_duplicate),
                status,
                reason,
                row.get("导出文件名", ""),
                row.get("创建时间", now) or now,
                now,
            ]
        )

        summary["total"] += 1
        summary[status] += 1
        per_shop[shop]["total"] += 1
        per_shop[shop][status] += 1
        if not buyer_id:
            summary["no_buyer_id"] += 1
            per_shop[shop]["no_buyer_id"] += 1
        if today_chatted:
            summary["today_chatted"] += 1
            per_shop[shop]["today_chatted"] += 1
        if has_unfinished_order:
            summary["unfinished"] += 1
            per_shop[shop]["unfinished"] += 1
        if is_duplicate:
            summary["duplicate"] += 1
            per_shop[shop]["duplicate"] += 1

    ws.freeze_panes = "A2"
    header_fill = PatternFill("solid", fgColor="00B050")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for column in ws.columns:
        letter = column[0].column_letter
        width = min(max(len(str(cell.value or "")) for cell in column) + 2, 42)
        ws.column_dimensions[letter].width = width

    ws_summary = wb.create_sheet("汇总")
    ws_summary.append(["店铺名", "总数", "待发送", "跳过", "异常-未找到旺旺ID", "今日已对话", "未完结订单", "当月重复"])
    for shop in SHOP_ORDER:
        data = per_shop.get(shop, {})
        ws_summary.append(
            [
                shop,
                data.get("total", 0),
                data.get("待发送", 0),
                data.get("跳过", 0),
                data.get("异常-未找到旺旺ID", 0),
                data.get("today_chatted", 0),
                data.get("unfinished", 0),
                data.get("duplicate", 0),
            ]
        )
    ws_summary.append(
        [
            "合计",
            summary.get("total", 0),
            summary.get("待发送", 0),
            summary.get("跳过", 0),
            summary.get("异常-未找到旺旺ID", 0),
            summary.get("today_chatted", 0),
            summary.get("unfinished", 0),
            summary.get("duplicate", 0),
        ]
    )
    for cell in ws_summary[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for column in ws_summary.columns:
        letter = column[0].column_letter
        width = min(max(len(str(cell.value or "")) for cell in column) + 2, 24)
        ws_summary.column_dimensions[letter].width = width

    wb.save(OUT_FILE)
    print(OUT_FILE)
    print(json.dumps({"summary": dict(summary), "per_shop": {k: dict(v) for k, v in per_shop.items()}}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
