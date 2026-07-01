from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from refund_recovery.send_results import SEND_HEADERS, SendResult, load_send_results

DEFAULT_WORKBOOK = ROOT / "exports" / "refund_followup_2026-06.xlsx"
DEFAULT_LOG = ROOT / "exports" / "send_test_log_2026-07-01.json"
DETAIL_SHEET = "跟进明细"


def normalize(value: object) -> str:
    return "" if value is None else str(value).strip()


def header_map(ws) -> dict[str, int]:
    return {normalize(cell.value): cell.column for cell in ws[1] if normalize(cell.value)}


def ensure_headers(ws, headers: list[str]) -> dict[str, int]:
    columns = header_map(ws)
    next_column = ws.max_column + 1
    for header in headers:
        if header not in columns:
            ws.cell(row=1, column=next_column, value=header)
            columns[header] = next_column
            next_column += 1
    return columns


def find_row(ws, columns: dict[str, int], result: SendResult) -> int | None:
    shop_col = columns["店铺名"]
    order_col = columns["订单号"]
    buyer_col = columns["旺旺ID"]
    for row_index in range(2, ws.max_row + 1):
        if normalize(ws.cell(row_index, shop_col).value) != result.shop:
            continue
        if normalize(ws.cell(row_index, order_col).value) != result.order_id:
            continue
        if result.buyer_id and normalize(ws.cell(row_index, buyer_col).value) != result.buyer_id:
            continue
        return row_index
    return None


def apply_result(ws, columns: dict[str, int], row_index: int, result: SendResult) -> None:
    ws.cell(row=row_index, column=columns["处理状态"], value=result.process_status)
    ws.cell(row=row_index, column=columns["更新时间"], value=result.sent_at)
    ws.cell(row=row_index, column=columns["发送状态"], value=result.send_status)
    ws.cell(row=row_index, column=columns["发送时间"], value=result.sent_at)
    ws.cell(row=row_index, column=columns["发送内容版本"], value=result.template_version)
    ws.cell(row=row_index, column=columns["发送详情"], value=result.detail)


def apply_send_log(workbook_path: Path, log_path: Path, output_path: Path | None = None) -> list[dict[str, str]]:
    results = load_send_results(json.loads(log_path.read_text(encoding="utf-8")))

    wb = load_workbook(workbook_path)
    ws = wb[DETAIL_SHEET] if DETAIL_SHEET in wb.sheetnames else wb.active
    columns = ensure_headers(ws, SEND_HEADERS)

    applied: list[dict[str, str]] = []
    for result in results:
        row_index = find_row(ws, columns, result)
        if row_index is None:
            applied.append({"order_id": result.order_id, "status": "not_found", "row": ""})
            continue
        apply_result(ws, columns, row_index, result)
        applied.append({"order_id": result.order_id, "status": result.send_status, "row": str(row_index)})

    wb.save(output_path or workbook_path)
    return applied


def main() -> None:
    parser = argparse.ArgumentParser(description="Apply QianNiu send logs to the monthly refund follow-up workbook.")
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK)
    parser.add_argument("--log", type=Path, default=DEFAULT_LOG)
    parser.add_argument("--output", type=Path)
    args = parser.parse_args()

    applied = apply_send_log(args.workbook, args.log, args.output)
    print(
        json.dumps(
            {"workbook": str(args.workbook), "output": str(args.output or args.workbook), "applied": applied},
            ensure_ascii=False,
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
